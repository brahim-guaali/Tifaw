from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)

TEXT_EXTENSIONS = {
    ".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".toml",
    ".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".css",
    ".go", ".rs", ".java", ".cpp", ".c", ".h", ".rb", ".sh",
    ".sql", ".xml", ".env", ".ini", ".cfg",
}

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}

PDF_EXTENSION = ".pdf"


@dataclass
class ExtractionResult:
    text_content: str | None = None
    image_bytes: bytes | None = None
    file_type: str = "unknown"
    metadata: dict | None = None


def extract_content(path: Path) -> ExtractionResult:
    ext = path.suffix.lower()

    if ext in IMAGE_EXTENSIONS:
        return _extract_image(path)
    elif ext == PDF_EXTENSION:
        return _extract_pdf(path)
    elif ext in TEXT_EXTENSIONS:
        return _extract_text(path)
    elif ext == ".docx":
        return _extract_docx(path)
    elif ext == ".xlsx":
        return _extract_xlsx(path)
    else:
        return ExtractionResult(file_type="binary")


def _extract_text(path: Path) -> ExtractionResult:
    try:
        text = path.read_text(errors="replace")[:2000]
        return ExtractionResult(text_content=text, file_type="text")
    except Exception as e:
        logger.warning("Failed to read text from %s: %s", path, e)
        return ExtractionResult(file_type="text")


def _extract_image(path: Path) -> ExtractionResult:
    try:
        image_bytes = path.read_bytes()
        metadata = _extract_image_metadata(path)
        return ExtractionResult(image_bytes=image_bytes, file_type="image", metadata=metadata)
    except Exception as e:
        logger.warning("Failed to read image %s: %s", path, e)
        return ExtractionResult(file_type="image")


def _extract_pdf(path: Path) -> ExtractionResult:
    text = ""
    image_bytes = None
    metadata = None

    try:
        import fitz  # pymupdf

        doc = fitz.open(str(path))

        # Extract text from all pages
        for page in doc:
            text += page.get_text() + "\n"
        text = text[:2000]

        # Render first page as image
        if len(doc) > 0:
            page = doc[0]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            image_bytes = pix.tobytes("png")

        # Extract PDF metadata
        metadata = _extract_pdf_metadata(doc)

        doc.close()
    except Exception as e:
        logger.warning("Failed to extract PDF %s: %s", path, e)

    return ExtractionResult(
        text_content=text or None, image_bytes=image_bytes, file_type="pdf", metadata=metadata
    )


def _extract_docx(path: Path) -> ExtractionResult:
    try:
        from docx import Document

        doc = Document(str(path))
        text = "\n".join(p.text for p in doc.paragraphs)[:2000]
        metadata = _extract_docx_metadata(doc)
        return ExtractionResult(text_content=text, file_type="docx", metadata=metadata)
    except ImportError:
        logger.info("python-docx not installed, skipping .docx extraction")
        return ExtractionResult(file_type="docx")
    except Exception as e:
        logger.warning("Failed to extract docx %s: %s", path, e)
        return ExtractionResult(file_type="docx")


def _extract_xlsx(path: Path) -> ExtractionResult:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        lines = []
        for sheet in wb.sheetnames[:3]:
            ws = wb[sheet]
            for row in ws.iter_rows(max_row=50, values_only=True):
                line = " | ".join(str(c) for c in row if c is not None)
                if line:
                    lines.append(line)
        metadata = _extract_xlsx_metadata(wb)
        wb.close()
        text = "\n".join(lines)[:2000]
        return ExtractionResult(text_content=text, file_type="xlsx", metadata=metadata)
    except ImportError:
        logger.info("openpyxl not installed, skipping .xlsx extraction")
        return ExtractionResult(file_type="xlsx")
    except Exception as e:
        logger.warning("Failed to extract xlsx %s: %s", path, e)
        return ExtractionResult(file_type="xlsx")


# --- Metadata extraction helpers ---


def _gps_to_decimal(gps_coords, gps_ref: str) -> float | None:
    """Convert EXIF GPS coordinates (degrees, minutes, seconds) to decimal."""
    try:
        degrees = float(gps_coords[0])
        minutes = float(gps_coords[1])
        seconds = float(gps_coords[2])
        decimal = degrees + minutes / 60 + seconds / 3600
        if gps_ref in ("S", "W"):
            decimal = -decimal
        return round(decimal, 6)
    except (TypeError, IndexError, ValueError, ZeroDivisionError):
        return None


def _extract_image_metadata(path: Path) -> dict | None:
    """Extract EXIF metadata from an image using Pillow."""
    try:
        from PIL import Image
        from PIL.ExifTags import TAGS, GPSTAGS

        img = Image.open(path)
        meta: dict = {
            "image_width": img.width,
            "image_height": img.height,
        }

        exif_data = img.getexif()
        if not exif_data:
            img.close()
            return meta

        # Basic EXIF tags
        tag_map = {
            271: "camera_make",      # Make
            272: "camera_model",     # Model
            274: "orientation",      # Orientation
            33434: "exposure_time",  # ExposureTime
            33437: "aperture",       # FNumber
            34855: "iso",            # ISOSpeedRatings
            36867: "date_taken",     # DateTimeOriginal
            36868: "date_digitized", # DateTimeDigitized
            37386: "focal_length",   # FocalLength
        }

        for tag_id, key in tag_map.items():
            val = exif_data.get(tag_id)
            if val is None:
                continue
            # Convert IFDRational to float
            if hasattr(val, "numerator") and hasattr(val, "denominator"):
                if key == "exposure_time" and val.denominator:
                    meta[key] = f"{val.numerator}/{val.denominator}"
                elif val.denominator:
                    meta[key] = round(float(val), 4)
            elif key == "date_taken" or key == "date_digitized":
                # Convert "2024:03:15 14:30:00" → "2024-03-15T14:30:00"
                try:
                    meta[key] = str(val).replace(":", "-", 2).replace(" ", "T", 1)
                except Exception:
                    meta[key] = str(val)
            else:
                meta[key] = val

        # GPS data (stored in IFD pointer tag 34853)
        gps_ifd = exif_data.get_ifd(0x8825)
        if gps_ifd:
            lat = gps_ifd.get(2)   # GPSLatitude
            lat_ref = gps_ifd.get(1)  # GPSLatitudeRef
            lon = gps_ifd.get(4)   # GPSLongitude
            lon_ref = gps_ifd.get(3)  # GPSLongitudeRef
            alt = gps_ifd.get(6)   # GPSAltitude

            if lat and lat_ref:
                meta["gps_latitude"] = _gps_to_decimal(lat, lat_ref)
            if lon and lon_ref:
                meta["gps_longitude"] = _gps_to_decimal(lon, lon_ref)
            if alt is not None:
                try:
                    meta["gps_altitude"] = round(float(alt), 1)
                except (TypeError, ValueError):
                    pass

        img.close()

        # Remove None values
        return {k: v for k, v in meta.items() if v is not None}
    except Exception as e:
        logger.debug("Failed to extract EXIF from %s: %s", path, e)
        return None


def _extract_pdf_metadata(doc) -> dict | None:
    """Extract metadata from an open PyMuPDF document."""
    try:
        raw = doc.metadata or {}
        meta = {}

        field_map = {
            "title": "title",
            "author": "author",
            "subject": "subject",
            "keywords": "keywords",
            "creator": "creator",
            "producer": "producer",
            "creationDate": "created",
            "modDate": "modified",
        }

        for src, dst in field_map.items():
            val = raw.get(src)
            if val:
                # PyMuPDF dates look like "D:20240315143000+00'00'"
                if src in ("creationDate", "modDate") and val.startswith("D:"):
                    val = _parse_pdf_date(val)
                meta[dst] = val

        meta["page_count"] = len(doc)
        return {k: v for k, v in meta.items() if v is not None}
    except Exception as e:
        logger.debug("Failed to extract PDF metadata: %s", e)
        return None


def _parse_pdf_date(date_str: str) -> str | None:
    """Parse PDF date format 'D:YYYYMMDDHHmmSS' to ISO format."""
    try:
        # Strip 'D:' prefix and timezone suffix
        s = date_str[2:16]  # YYYYMMDDHHmmSS
        if len(s) >= 8:
            year, month, day = s[0:4], s[4:6], s[6:8]
            hour = s[8:10] if len(s) >= 10 else "00"
            minute = s[10:12] if len(s) >= 12 else "00"
            second = s[12:14] if len(s) >= 14 else "00"
            return f"{year}-{month}-{day}T{hour}:{minute}:{second}"
    except Exception:
        pass
    return None


def _extract_docx_metadata(doc) -> dict | None:
    """Extract core properties from a python-docx Document."""
    try:
        props = doc.core_properties
        meta = {}

        if props.author:
            meta["author"] = props.author
        if props.title:
            meta["title"] = props.title
        if props.subject:
            meta["subject"] = props.subject
        if props.keywords:
            meta["keywords"] = props.keywords
        if props.category:
            meta["category"] = props.category
        if props.last_modified_by:
            meta["last_modified_by"] = props.last_modified_by
        if props.revision is not None:
            meta["revision"] = props.revision
        if props.created:
            meta["created"] = props.created.isoformat()
        if props.modified:
            meta["modified"] = props.modified.isoformat()

        return meta if meta else None
    except Exception as e:
        logger.debug("Failed to extract DOCX metadata: %s", e)
        return None


def _extract_xlsx_metadata(wb) -> dict | None:
    """Extract properties from an openpyxl workbook."""
    try:
        props = wb.properties
        meta = {}

        if props.creator:
            meta["author"] = props.creator
        if props.title:
            meta["title"] = props.title
        if props.created:
            meta["created"] = props.created.isoformat()
        if props.modified:
            meta["modified"] = props.modified.isoformat()
        meta["sheet_count"] = len(wb.sheetnames)

        return meta if meta else None
    except Exception as e:
        logger.debug("Failed to extract XLSX metadata: %s", e)
        return None
